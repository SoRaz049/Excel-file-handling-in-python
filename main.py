from tkinter.tix import Y_REGION
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

acc_list=[]
pp_bal=[]

wb= load_workbook ("D:\\Projects\\Python\\Excel file handling\\account_sheet.xlsx")

ws= wb.active


for col in range (2,3):
    for row in range (2,11):
        char = get_column_letter( col )
        
        acc_no= ws [char + str(row)].value
        acc_list.append(acc_no)
        

print(acc_list)
        
sender=int (input("Enter the account number of the sender: "))
receiver=int (input("Enter the account number of the reciptant: "))

sen_index=(acc_list.index(str (sender)))
rec_index=(acc_list.index(str (receiver)))


# for col in range (3,4):
#     for row in range (2,11):
#         char=get_column_letter(col)
#         p_balance= ws [char + str(row)].value
#         pp_bal.append(p_balance)

send_amount = int (input("Enter how much money you want to tansfer? "))
sender_previous = int(ws['C'+ str(sen_index + 2)].value) 
receiver_previous = int(ws['C'+ str(rec_index + 2)].value) 

ws['C'+ str(sen_index + 2)].value = sender_previous - send_amount
ws['C'+ str(rec_index + 2)].value = receiver_previous + send_amount
print("vayo yr")
print(ws['C'+ str(rec_index + 2)].value)

wb.save("D:\\Projects\\Python\\Excel file handling\\account_sheet.xlsx")

